import os
import xlrd
from openpyxl import load_workbook
from typing import List, Dict, Any


class ExcelReaderUtil:
    """
    Python Excel通用读取工具类
    支持 .xls (xlrd) 和 .xlsx (openpyxl) 格式
    """

    @staticmethod
    def read_excel(file_path: str, sheet_name: str = None, sheet_index: int = 0, skip_header: bool = True) -> List[Dict[str, Any]]:
        """
        读取Excel文件数据
        :param file_path: Excel文件路径（绝对/相对路径）
        :param sheet_name: 要读取的sheet名称（优先级高于sheet_index）
        :param sheet_index: 要读取的sheet索引（从0开始），仅当sheet_name为None时生效
        :param skip_header: 是否跳过表头（第一行）
        :return: 列表，每个元素是字典（key=列名/列索引，value=单元格值）
        """
        # 校验文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel文件不存在：{file_path}")

        # 识别文件格式并调用对应读取方法
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == ".xlsx":
            return ExcelReaderUtil._read_xlsx(file_path, sheet_name, sheet_index, skip_header)
        elif file_ext == ".xls":
            return ExcelReaderUtil._read_xls(file_path, sheet_name, sheet_index, skip_header)
        else:
            raise ValueError(f"不支持的文件格式：{file_ext}，仅支持.xls和.xlsx")

    @staticmethod
    def _read_xlsx(file_path: str, sheet_name: str, sheet_index: int, skip_header: bool) -> List[Dict[str, Any]]:
        """读取.xlsx格式文件"""
        # 加载工作簿（只读模式提升性能）
        wb = load_workbook(file_path, read_only=True, data_only=True)
        # 获取目标sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[sheet_index]

        # 解析数据
        result = []
        # 获取表头（列名）
        headers = []
        start_row = 1  # openpyxl的行索引从1开始
        if skip_header:
            header_row = ws[1]
            headers = [cell.value for cell in header_row if cell.value is not None]
            start_row = 2

        # 遍历数据行
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if all(cell is None for cell in row):  # 跳过空行
                continue
            # 构建行数据字典
            row_data = {}
            for idx, cell_value in enumerate(row):
                # 有表头则用表头作为key，无表头则用列索引作为key
                key = headers[idx] if (skip_header and idx < len(headers)) else f"col_{idx}"
                row_data[key] = cell_value
            result.append(row_data)

        wb.close()
        return result

    @staticmethod
    def _read_xls(file_path: str, sheet_name: str, sheet_index: int, skip_header: bool) -> List[Dict[str, Any]]:
        """读取.xls格式文件"""
        # 加载工作簿
        wb = xlrd.open_workbook(file_path)
        # 获取目标sheet
        if sheet_name and sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(sheet_index)

        # 解析数据
        result = []
        headers = []
        start_row = 0  # xlrd的行索引从0开始
        if skip_header:
            header_row = ws.row_values(0)
            headers = [cell for cell in header_row if cell is not None]
            start_row = 1

        # 遍历数据行
        for row_num in range(start_row, ws.nrows):
            row = ws.row_values(row_num)
            if all(cell is None or cell == "" for cell in row):  # 跳过空行
                continue
            # 构建行数据字典
            row_data = {}
            for idx, cell_value in enumerate(row):
                key = headers[idx] if (skip_header and idx < len(headers)) else f"col_{idx}"
                row_data[key] = cell_value
            result.append(row_data)

        return result


